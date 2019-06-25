#!/usr/bin/python
# -*- coding: utf-8 -*-
import os
import re
import ConfigParser,sys
import magic
import openpyxl
from openpyxl.styles import Font
# import zipfile
# import shutil
# import time, itertools

from Patternset import Patternset
# from KeywordPattern import KeywordPattern

class Keywordsgrepper:

    def __init__(self):
        self.patternset = Patternset()
        self.defaultrootpath = ''
        self.filepath = ''
        self.outputlog = {}
        self.workbook = openpyxl.Workbook()
        self.reportMatrix = {}
        self.init_ReportValue()


    def init_ReportValue(self):
        keys = self.patternset.keywordpattern.keys()
        for key in keys:
            self.reportMatrix[key] = 'False'


    # read Setup config
    def init_Setup(self):
        folderList = []  #record FWlog list location and FWlog data type
        setupFile = '' # use .ini file to read inputRoute and outputRoute
        try:
            setupFile = sys.argv[1]
        except:
            setupFile = input(">>> SetupPath: ")

        if(os.path.exists(setupFile)):
            # read setup config ex. setup.ini
            setup = ConfigParser.ConfigParser()
            setup.read(setupFile)
            # Setup output path
            outputRoute = setup.get('Output Path','o')
            # Setup input path
            pathKeys = setup.options('Read Path')
            for pathKey in pathKeys :
                if(os.path.exists(setup.get('Read Path',pathKey))):
                    folderList.append(setup.get('Read Path',pathKey))
                    self.defaultrootpath = setup.get('Read Path',pathKey)
        else:
            print('Error:init_setup Failed. \nReason: does not exist the setup file.')
            exit()

        return folderList,outputRoute


    # 掃描目前資料夾中有哪些檔
    # iterate Scan how many file in input path
    def getFileName(self,route):
        csvList = []

        for root, dirs, files in os.walk(route):
            for file in files:
                csvList.append(os.path.join(root, file))
                # input(os.path.join(root, file))
                # extension = os.path.splitext(os.path.join(root, file))[-1]
                # input(extension)
        # raw_input(len(csvList))

        return csvList

    """
    ----------------------------------------------------
    -----------------get File MIME Type-----------------
    ----------------------------------------------------
    """
    def getFileInfo(self,filepath):
        # whitelist = ['Apple Desktop Services Store','POSIX tar archive (GNU)','POSIX','GNU','UBI image']
        file_type = magic.from_file(filepath)
        mime_type = magic.from_file(filepath,mime=True)

        return file_type,mime_type


    """
    ----------------------------------------------------
    -----------------Main Grepper Function--------------
    ----------------------------------------------------
    """
    def grepCenter(self,file_path):
        try:
            lines = self.readDataCenter(file_path)
            patternset = Patternset()
            temp_output = {}

            if getattr(lines, '__iter__', None):

                for index, line in enumerate(lines):

                    for key in patternset.keywordpattern:
                        # use regular expression to grep keywords
                        match_objlist = re.findall(patternset.keywordpattern[key], line)

                        if match_objlist:
                            for item in match_objlist:
                                #check ipv4
                                if 'ipv4' in key :
                                    file_name, file_extension = os.path.splitext(file_path)
                                    # raw_input(file_extension)
                                    if 'crt' in file_extension:
                                        break
                                    if self.isPublicIPaddress(item) :
                                        #First time init
                                        if key not in temp_output.keys():
                                            temp_output[key] = [item]
                                        elif item not in temp_output[key]:
                                            temp_output[key].append(item)
                                            # print("Keyword: %s Context: %s filepath: %s"%(key,item,file_path))

                                else :
                                    if key not in temp_output.keys():
                                        temp_output[key] = [item]
                                    elif item not in temp_output[key]:
                                        temp_output[key].append(item)
                                        # print("Keyword: %s Context: %s filepath: %s" % (key, item, file_path))
                        else :
                            continue

            return temp_output

        except:
            print('have Some Error !')

    def absolutepath2relativepath(self):
        for sheetname in self.workbook.sheetnames:
            for row in self.workbook[sheetname].iter_rows():
                for cell in row:

                    # change absolute filepath to relative path
                    if cell.column == 'A' and cell.row > 1:
                        absolute_path = str(cell.value)
                        relative_path = absolute_path.replace(self.defaultrootpath, '.')
                        cell.value = relative_path

    def change_excel_Style(self):
        red_font = Font(color='00FF0000', size=14)
        normal_font = Font(size=14)

        for sheetname in self.workbook.sheetnames:
            # change Font size and color
            for row in self.workbook[sheetname].iter_rows():
                for cell in row:
                    if cell.value == 'True':
                        cell.font = red_font
                    else:
                        cell.font = normal_font

            # change cell width
            for col in self.workbook[sheetname].columns:
                max_length = 0
                column = col[0].column  # Get the column name
                for cell in col:
                    try:  # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                self.workbook[sheetname].column_dimensions[column].width = adjusted_width

    def generate_Report(self):

        default_sheet = self.workbook.get_sheet_by_name('Sheet')
        if default_sheet:
            default_sheet.title = 'TotalReport'
        else:
            self.workbook.create_sheet(title='TotalReport')

        first_row = self.reportMatrix.keys()
        first_row.insert(0,'Test Module:')

        self.workbook['TotalReport'].append(first_row)

        keys = self.reportMatrix.keys()
        second_row = ['Result:']
        for key in keys:
            second_row.append(self.reportMatrix[key])
        self.workbook['TotalReport'].append(second_row)

        # for key in keys:


    # write keyword into excel file
    def write_xlsx(self,filepath,data):
        for key in data.keys():
            # raw_input(key)
            self.reportMatrix[key] = 'True'

            new_data_list = []


            if key not in self.workbook.sheetnames:
                self.workbook.create_sheet(title=key)
                self.workbook[key].append(['FilePath','Data'])
            max_row = self.workbook[key].max_row
            # raw_input(max_row)

            items = data[key]
            for item in items:
                self.workbook[key].append([filepath,str(item)])


    def isPublicIPaddress(self,ip):
        # raw_input(ip.split('.'))

        ip_segment = map(int,ip.split('.'))

        # check ipv4 format is legal
        for item in ip_segment:
            if item > 255:
                return False

        # check local host ip address
        if re.search('(127.0.0.1)|(^0.0.0.0)|(255.255.255.0)|(255.255.255.255)',ip):
            return False
        #check is private ip address
        elif ip_segment[0] == 10 or ( ip_segment[0] ==172 and ip_segment[1] >=16 and ip_segment[1] <= 31) or (ip_segment[0] == 192 and ip_segment[1] == 168):
            return False
        # public ip address validation success
        else :
            return True

    def readDataCenter(self,file_path):
        logdata = self.openlogfile(file_path)
        # if file_extension == '.log' or file_extension == '.csv' or file_extension == '.txt' or file_extension == '':
        #     logdata = openlogfile(file_path)
        #     return logdata
        # elif file_extension == '.zip':
        #     zipdata = openzipfile(file_path)
        #     return zipdata
        # elif file_extension == '.gz':
        #     gzdata = opengzfile(file_path)
        #     return gzdata
        # elif file_extension == '.bz2':
        #     bz2data = openbz2file(file_path)
        #     return bz2data
        #     # elif file_extension == '.7z':
        #     #     zipdata = open7zfile(file_path)
        #     #     return zipdata
        # else:
        #     return ''
        return logdata



    def openlogfile(self,file_path):
        if os.path.exists(file_path):
            with open(file_path, 'rb') as SSG550:
                for SSG550Reader in SSG550:
                    yield SSG550Reader.decode('utf-8', 'ignore')
                    # input('asd')
        else:
            print('Check your log or csv file path is exist!')

    def save(self,outputpath):
        # raw_input(self.workbook.sheetnames)
        if not os.path.exists(outputpath):
            os.makedirs(outputpath)
        outputfile = outputpath+'/IoT_Firmware_Keyword_Detection_Report.xlsx'
        self.workbook.save(outputfile)
        print("Saved Report to Output Path :"+outputfile)

    # check file is important or not
    def check_file(self,patternset,filepath):
        temp_output = {}

        for key in patternset.pathpattern:
            match_obj = re.search(patternset.pathpattern[key],filepath)
            if match_obj:
                # raw_input(match_obj.group())
                temp_output['SensitiveFile'] = [match_obj.group()]
                return temp_output

        return None

"""
----------------------------------------------------
-----------------Main Function----------------------
----------------------------------------------------
"""

if '__main__' == __name__ :
    # init Object
    keywordsgrepper = Keywordsgrepper()

    print "=====KeywordsGrepper Tools v1.0======"


    if len(sys.argv) < 2:
        print("missing setup.ini argument!!!")
        exit()

    path = sys.argv[1]
    inputFolderPath,OutputPath = keywordsgrepper.init_Setup()

    print("Input Folder Path:"+inputFolderPath[0])

    file_List = keywordsgrepper.getFileName(inputFolderPath[0])

    patternset = Patternset()

    for file_path in file_List:
        # Search Sensitive File
        output = keywordsgrepper.check_file(patternset,file_path)
        if output:
            # raw_input(sensitivefile)
            keywordsgrepper.write_xlsx(file_path,output)

        # Search Sensitive Strings

        if os.path.isfile(file_path):
            file_name, file_extension = os.path.splitext(file_path)
            file_info = keywordsgrepper.getFileInfo(file_path)

            # if file_type is plaintext file ,then let's try to find some keyword in file
            mimetype = file_info[1]

            if ('text' in mimetype):
                output = keywordsgrepper.grepCenter(file_path)
                if output:
                    keywordsgrepper.write_xlsx(file_path,output)

        else:
            #file is broken symbolic link
            continue

    # Generate Excel format (.xlsx) Report to output folder
    keywordsgrepper.generate_Report()
    keywordsgrepper.change_excel_Style()
    keywordsgrepper.absolutepath2relativepath()
    keywordsgrepper.save(OutputPath)
    print("Finish~")


